#!/usr/bin/env python2
# -*- coding: utf-8 -*-
"""
Created on Sun Jul  9 12:54:23 2017

@author: salama
"""
import numpy as np
import tensorflow as tf
import joblib
from openpyxl import load_workbook,Workbook
import scipy.io as sio
import preparing_images_training
import time

faces_data = sio.loadmat("faces_data_colors_224.mat")
x_images = faces_data['images']
y_labels = faces_data['labels']
y_labels = np.reshape(y_labels,y_labels.shape[1])

#80% training , 20% test
tr= int(0.8 * y_labels.shape[0])
#ts= y_labels.shape[0] - tr

xtr = x_images[0:tr,:]
ytr = y_labels[0:tr]
xts = x_images[tr: , :]
yts = y_labels[tr: ]

pretrained_file = 'squeezenet_v1.1.pkl'



class SqueezeDet_Model(object):
    def __init__(self):
        #TODO: change hight,width,channels of input_images
        self.batch_size = 50
        self.image_hight = 224 #224
        self.image_width = 224 #224
        self.image_channels = 3
        
        self.input_images = tf.placeholder(tf.float32,
                                           [None,self.image_hight,self.image_width,self.image_channels],
                                           name='image_batch_input')

        
        self.pretrained_model = joblib.load(pretrained_file)

        self.classes = 5
        
        
        self.labels = tf.placeholder(tf.int64,[None])
        
        self.max_steps = 1000000
        self.lr = 0.00001
        
        
        #y_test = tf.placeholder(tf.int32,[None,n_classes])
        self.keep_prob = tf.placeholder(tf.float32)

    
    def conv_layer(self,input_activation,size,stride,filters,layer_name = None,
                   freeze = False,wd = None, pretrained = False,
                   padding = 'SAME',relu = False):
        with tf.variable_scope(layer_name):
            channels = np.shape(input_activation)[-1]
            '''
            a__ = np.array(input_activation)
            print(type(a__))
            print(a__.shape)
            channels = a__.shape[-1]
            #print(input_activation)
            '''
            if pretrained :
                # read caffe kernel weights [out,in,hight,width]
                kernels_values = self.pretrained_model[layer_name][0]
                # convert to tensorflow kernel weights [hight,width,in,out]
                kernels_values = np.transpose(kernels_values,[2,3,1,0])
                bias_values = self.pretrained_model[layer_name][1]
                
                # add initialized tf.variable and weight decay(regularization)
                kernels_init = tf.constant_initializer(kernels_values)
                bias_init = tf.constant_initializer(bias_values)
            else:
                # initialize random kerenel wieghts and bias
                kernels_init = tf.truncated_normal_initializer()
                bias_init = tf.constant_initializer(0.0)
                
            # define variable weights to be trained
            kernels = tf.get_variable('Weights',shape=[int(size),int(size),int(channels),int(filters)],
                                      initializer = kernels_init,
                                      trainable = not freeze)
            bias = tf.get_variable('Bias',shape=[filters],
                                   initializer = bias_init,
                                   trainable = not freeze)
            if wd is not None and (not freeze) :
                # regularization
                weight_decay = tf.multiply(tf.nn.l2_loss(kernels),wd)
                # TODO: add weight_decay to the total loss function
                
            
            # make the convolution operation with bias 
            out = tf.nn.conv2d(input_activation,
                               kernels,
                               [1,stride,stride,1],
                               padding = padding,
                               name ='conv')
            out = tf.nn.bias_add(out,bias,name ='add_bias')
            
            if relu:
                out = tf.nn.relu(out)
            return out
    
    def max_pool_layer(self,input_activation,size,stride,
                       padding = 'SAME',layer_name = None):
        with tf.variable_scope(layer_name):
            out = tf.nn.max_pool(input_activation,[1,size,size,1],
                                 [1,stride,stride,1],padding)
            return out
    
    def fire_module(self,input_activation,s1x1,e1x1,e3x3,
                    layer_name = None,pretrained = False,freeze = False):
        
        with tf.variable_scope(layer_name):
            out1 = self.conv_layer(input_activation,1,1,s1x1, # input, size, stride, filters
                              layer_name=layer_name+'/squeeze1x1',pretrained=pretrained,
                              padding='SAME', freeze=freeze)
            out2 = self.conv_layer(out1,1,1,e1x1, # input, size, stride, filters
                              layer_name=layer_name+'/expand1x1',pretrained=pretrained,
                              padding='SAME', freeze=freeze)
            out3 = self.conv_layer(out1,3,1,e3x3, # input, size, stride, filters
                              layer_name=layer_name+'/expand3x3',pretrained=pretrained,
                              padding='SAME', freeze=freeze)
            return tf.concat([out2, out3], 3)
        
        
    def layers(self):
        '''this function for define the layers structure of the neural network
        as a stack of conv and maxpool layers acording to the ConvDet proposed
        paper.
        hint : last layer can be modified to be suitable to the dataset (number
        of classes), and suitable for the porposed model as number of anchors'''
        # from 1 to 9 all pretrained model
        conv1 = self.conv_layer(self.input_images,3,2,64,
                                layer_name='conv1',padding='VALID',
                                pretrained=True,relu=True,freeze=True)
        maxpool1 = self.max_pool_layer(conv1,3,2,layer_name='maxpool1',padding='VALID')
        fire2 = self.fire_module(maxpool1,16,64,64,layer_name='fire2',pretrained=True)
        fire3 = self.fire_module(fire2,16,64,64,layer_name='fire3',pretrained=True)
        maxpool3 = self.max_pool_layer(fire3,3,2,layer_name='maxpool3',padding='VALID')
        fire4 = self.fire_module(maxpool3,32,128,128,layer_name='fire4',pretrained=True)
        fire5 = self.fire_module(fire4,32,128,128,layer_name='fire5',pretrained=True)
        maxpool5 = self.max_pool_layer(fire5,3,2,layer_name='maxpool5',padding='VALID')
        fire6 = self.fire_module(maxpool5,48,192,192,layer_name='fire6',pretrained=True)
        fire7 = self.fire_module(fire6,48,192,192,layer_name='fire7',pretrained=True)
        fire8 = self.fire_module(fire7,64,256,256,layer_name='fire8',pretrained=True)
        fire9 = self.fire_module(fire8,64,256,256,layer_name='fire9',pretrained=True)
        
        # new layers (not pretrained)
        fire10 = self.fire_module(fire9,96,384,384,layer_name='fire10',pretrained=False)
        fire11 = self.fire_module(fire10,96,384,384,layer_name='fire11',pretrained=False)
        # TODO: make keep_prob as input to the run operation
        dropout11 = tf.nn.dropout(fire11,keep_prob=self.keep_prob,name='dropout11') 
        # TODO: make numper of filters = num predictions = # anchors per grid * (# classes + 1confidence_prob + 4bbox)
        # for voc dataset -> # filters = 225
        convDet = self.conv_layer(dropout11,13,1,self.classes,
                                layer_name='convDet',padding='VALID',
                                pretrained=False,relu=False)
	
        convDet = tf.reshape(convDet,[-1,self.classes])
        self.preds = convDet
        return convDet
    
    
    def train_neural_network(self):
        batch_train = preparing_images_training.Batch(xtr,ytr,self.batch_size) 
	     
        prediction = self.preds
        cost = tf.reduce_mean( tf.nn.sparse_softmax_cross_entropy_with_logits(logits=prediction, labels=self.labels) )
        optimizer = tf.train.AdamOptimizer(learning_rate=self.lr,beta1=0.9).minimize(cost)
        
        hm_epochs = 1
        with tf.Session() as sess:
            sess.run(tf.global_variables_initializer())
            
            writer =  tf.summary.FileWriter('./graphs',sess.graph)
            
            correct = tf.equal(tf.argmax(prediction, 1), self.labels)
            accuracy = tf.reduce_mean(tf.cast(correct, 'float'))
            saver = tf.train.Saver()
            #saver_w = tf.train.Saver(weights)
            #saver_b = tf.train.Saver(biases)
            global checkpoint
            
            
            if not(checkpoint==0):
                saver.restore(sess, file_path)
                book = load_workbook("FaceDet_SqueezeNet.xlsx")
                sheet1 = book.get_sheet_by_name("Sheet1")
            else :
                book = Workbook()
                sheet1 = book.create_sheet("Sheet1") #.add_sheet('Sheet1',cell_overwrite_ok=True)
                book.save('FaceDet_SqueezeNet.xlsx')
            
            
            for epoch in range(hm_epochs):
                epoch_loss = 0
                start_time = time.time()
                
                
                
                for _ in range(int(xtr.shape[0]/self.batch_size)):
                    epoch_x, epoch_y = batch_train.getBatch_series()   # TODO: 
                    _, c = sess.run([optimizer, cost], feed_dict={self.input_images: epoch_x, self.labels: epoch_y ,self.keep_prob: 0.5})
                    epoch_loss += c
                
                duration = time.time() - start_time
                print("Epoch lasts for : ",duration)
                print('Epoch', epoch, 'completed out of',hm_epochs,'loss:',epoch_loss)
                # save a checkpoint
                saver.save(sess,file_path)
                #saver_w.save(sess,weights_path)
                #saver_b.save(sess,biases_path)
                
                checkpoint = 1
                print("weights have been saved")
                # test against the validation data set
                '''
		percent = 0
                for batch in range(int(validation.num_examples/batch_size)):
                    val_x, val_y = validation.next_batch(batch_size)
                    percent += accuracy.eval({x:val_x, y:val_y, keep_prob: 1})
                    #print percent
                percent /= (validation.num_examples/batch_size)
                print('Validation Accuracy:', percent)
                sheet1.append({"A":percent})
                book.save('FaceDet_SqueezeNet.xlsx')    
                '''
                if(epoch%10==9 or epoch==hm_epochs-1):
                    # check accuracy aginst test data
                    percent = 0
                    batch_test = preparing_images_training.Batch(xts,yts,20) 
                    for batch in range(int(xts.shape[0]/20)):
                        #correct = tf.equal(tf.argmax(prediction, 1), tf.argmax(y, 1))
                        #accuracy = tf.reduce_mean(tf.cast(correct, 'float'))
                        test_x, test_y = batch_test.getBatch_series()  # TODO : some thing wrong
                        percent += accuracy.eval({self.input_images:test_x, self.labels:test_y, self.keep_prob: 1})
                    print('Accuracy:',percent/(int(xts.shape[0]/20)))
                    


checkpoint = 0
file_path = "D:/4th year 2nd term/GP/Facedet_Final_version/facedet.ckpt"
#weights_path = "/home/ti1080/GP_TEAM2_salama/Facedet_v2/weights/facedet_weights.ckpt"
#biases_path = "/home/ti1080/GP_TEAM2_salama/Facedet_v2/weights/facedet_biases.ckpt"

mm = SqueezeDet_Model()
mm.layers()
mm.train_neural_network()
