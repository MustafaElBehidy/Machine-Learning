#!/usr/bin/env python2
# -*- coding: utf-8 -*-
"""
Created on Sat May 20 21:59:23 2017

@author: salama
"""

from __future__ import absolute_import
import tensorflow as tf
import numpy as np
import time
from openpyxl import load_workbook,Workbook
import scipy.io as sio
import preparing_images_training


def make_one_hot(labels,num_classes):
    num_examples = labels.shape[0]
    one_hot = np.reshape([False]* num_examples *num_classes,[num_examples,num_classes])
    one_hot[range(num_examples),labels] = True
    return one_hot

batch_size = 50
classes = 5 
image_height = 224 
image_width =  224
image_channels = 3
image_size = image_height * image_width *image_channels
learning_rate = 0.001

faces_data = sio.loadmat("faces_data_colors_224.mat")
x_images = faces_data['images']
y_labels = faces_data['labels']
y_labels = np.reshape(y_labels,y_labels.shape[1])

x = tf.placeholder(tf.float32, [None, image_height* image_width *image_channels ] )
y = tf.placeholder(tf.float32, [None ,classes])
#y = tf.placeholder(tf.int32, [None])
keep_prob = tf.placeholder(tf.float32)

#80% training , 20% test
tr= int(0.8 * y_labels.shape[0])
#ts= y_labels.shape[0] - tr

xtr = x_images[0:tr,:]
ytr = y_labels[0:tr]
ytr = make_one_hot(ytr,classes)
xts = x_images[tr: , :]
yts = y_labels[tr: ]
yts = make_one_hot(yts,classes)



def conv2d(x, W, name=None):
  return tf.nn.conv2d(x, W, strides=[1, 1, 1, 1], padding='SAME', name=name)

def maxpool2d(x, name=None):
  return tf.nn.max_pool(x, ksize=[1, 2, 2, 1], strides=[1, 2, 2, 1], padding='SAME', name=name)

def convolutional_neural_network(x):#, keep_rate):
    weights = {
        
        'W_conv1': tf.Variable(tf.random_normal([3, 3, image_channels , 16],stddev=0.01),name='conv1_w'),
        
        'W_conv2': tf.Variable(tf.random_normal([3, 3, 16, 32],stddev=0.01),name='conv2_w'),
        
        'W_conv3': tf.Variable(tf.random_normal([3, 3, 32, 32],stddev=0.01),name='conv3_w'),
                                                
        'W_conv4': tf.Variable(tf.random_normal([3, 3, 32, 64],stddev=0.01),name='conv4_w'),
                                                
        'W_conv5': tf.Variable(tf.random_normal([3, 3, 64, 64],stddev=0.01),name='conv5_w'),                                        
        
        'W_fc1': tf.Variable(tf.random_normal([28*28*64, 512],stddev=0.01),name='fc1_w'),
                                              
        'W_fc2': tf.Variable(tf.random_normal([512, 128],stddev=0.01),name='fc2_w'),
        
        'out': tf.Variable(tf.random_normal([128, classes],stddev=0.01),name='out_w')
    }
    
    biases = {
        'b_conv1': tf.Variable(tf.ones([16]),name='conv1_b'),
        'b_conv2': tf.Variable(tf.ones([32]),name='conv2_b'),
        'b_conv3': tf.Variable(tf.ones([32]),name='conv3_b'),
        'b_conv4': tf.Variable(tf.ones([64]),name='conv4_b'),
        'b_conv5': tf.Variable(tf.ones([64]),name='conv5_b'),
                                       
        'b_fc1': tf.Variable(tf.ones([512]),name='fc1_b'),
        'b_fc2': tf.Variable(tf.ones([128]),name='fc2_b'),
        'out': tf.Variable(tf.ones([classes]),name='out_b')
    }
    
    global saver_w
    global saver_b
    saver_w = tf.train.Saver(weights)
    saver_b = tf.train.Saver(biases)
    
    # Reshape input to a 4D tensor 
    x = tf.reshape(x, shape=[-1, image_height , image_width , image_channels ])
    
    conv1 = tf.nn.relu(conv2d(x, weights['W_conv1'],name='CONV1') + biases['b_conv1'])
    conv2 = tf.nn.relu(conv2d(conv1, weights['W_conv2'],name='CONV2') + biases['b_conv2'])
    conv2 = maxpool2d(conv2, name='MAX_POOL1')
    conv2 = tf.nn.dropout(conv2, keep_prob, name='DROPOUT1')
    
    conv3 = tf.nn.relu(conv2d(conv2, weights['W_conv3'],name='CONV3') + biases['b_conv3'])
    conv4 = tf.nn.relu(conv2d(conv3, weights['W_conv4'],name='CONV4') + biases['b_conv4'])
    conv4 = maxpool2d(conv4, name='MAX_POOL2')
    conv4 = tf.nn.dropout(conv4, keep_prob, name='DROPOUT2')
    
    conv5 = tf.nn.relu(conv2d(conv4, weights['W_conv5'],name='CONV5') + biases['b_conv5'])
    conv5 = maxpool2d(conv5, name='MAX_POOL2')
    
    # Fully connected layer
    # Reshape conv2 output to fit fully connected layer
    fc = tf.reshape(conv5, [-1, 28*28*64])
    fc1 = tf.nn.relu(tf.matmul(fc, weights['W_fc1'],name='FC1') + biases['b_fc1'])
    fc1 = tf.nn.dropout(fc1, keep_prob, name='DROPOUT')
    
    fc2 = tf.nn.relu(tf.matmul(fc1, weights['W_fc2'],name='FC2') + biases['b_fc2'])
    
    output = tf.matmul(fc2, weights['out'], name='OUT') + biases['out']

    return output



    
def train_neural_network(x):
    prediction = convolutional_neural_network(x)
    cost = tf.reduce_mean( tf.nn.softmax_cross_entropy_with_logits(logits=prediction, labels=y) )
    optimizer = tf.train.AdamOptimizer(learning_rate=learning_rate,beta1=0.9).minimize(cost)
    
    hm_epochs = 10
    with tf.Session() as sess:
        sess.run(tf.global_variables_initializer())
        
        writer =  tf.summary.FileWriter('./graphs',sess.graph)
        
        correct = tf.equal(tf.argmax(prediction, 1), tf.argmax(y, 1))
        accuracy = tf.reduce_mean(tf.cast(correct, 'float'))
        saver = tf.train.Saver()
        
        global checkpoint
        
        
        if not(checkpoint==0):
            saver.restore(sess, file_path)
            book = load_workbook("facedet.xlsx")
            sheet1 = book.get_sheet_by_name("sheet1")
        else :
            book = Workbook()
            sheet1 = book.create_sheet("sheet1") #.add_sheet('Sheet1',cell_overwrite_ok=True)
            book.save('facedet.xlsx')
        
        #print (prediction.eval({x:np.ones((50,224*224*3))}))
        for epoch in range(hm_epochs):

            epoch_loss = 0
            start_time = time.time()
            batch_train = preparing_images_training.Batch(xtr,ytr,batch_size)
            for _ in range(int(xtr.shape[0]/batch_size)):
                epoch_x, epoch_y = batch_train.getBatch_series() 
                _, c = sess.run([optimizer, cost], feed_dict={x: epoch_x, y: epoch_y ,keep_prob: 0.5})
                epoch_loss += c
            
            duration = time.time() - start_time
            print("Epoch lasts for : ",duration)
            print('Epoch', epoch, 'completed out of',hm_epochs,'loss:',epoch_loss)
            
            saver.save(sess,file_path)
            
            
            checkpoint = 1
            print("weights have been saved")
            
            if(epoch or epoch == 0 or epoch%10==9 or epoch==hm_epochs-1):
                percent = 0
                batch_test = preparing_images_training.Batch(xts,yts,20)
                for batch in range(int(xts.shape[0]/20)):
                    test_x, test_y = batch_test.getBatch_series()
                    #print(test_y.shape)
                    percent += accuracy.eval({x:test_x, y:test_y, keep_prob: 1})
                    
                print('Accuracy:',percent/(int(xts.shape[0]/20)))

checkpoint = 1
file_path = "/home/ti1080/GP_TEAM2_salama/Facedet_v2/facedet.ckpt"
weights_path = "/home/ti1080/GP_TEAM2_salama/Facedet_v2/weights/facedet_weights.ckpt"
biases_path = "/home/ti1080/GP_TEAM2_salama/Facedet_v2/weights/facedet_biases.ckpt"


train_neural_network(x)
